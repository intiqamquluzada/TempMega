import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from django.conf import settings
from django.core.files.storage import FileSystemStorage

output_dir = "/home/Mega_App/media/"
#os.makedirs(output_dir, exist_ok=True)

def resulter(formatted_date, form1_a, form1_b, form2_a, form2_b,
             form_2yekun, form_4, form_4yekun,
             form_5, form_5yekun, form_6, form_7, form_10,
             form_3, form_8, form_9, form_11, form_12, form_13, class_of_insurance):
    def write_headers(ws, row_num, headers, data):
        for row_data in headers:
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=len(data.columns) + 1)
            cell = ws.cell(row=row_num, column=2, value=row_data[0])
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            row_num += 1
        return row_num

    def add_table(ws, row_num, columns, data):
        ws.cell(row=row_num, column=1, value='A')
        fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        for col_num, col_data in enumerate(columns, 2):
            cell = ws.cell(row=row_num, column=col_num, value=col_data)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = fill

        for index, row in data.iterrows():
            row_num += 1
            for col_num, value in enumerate(row, 2):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return row_num + 1

    wb = openpyxl.Workbook()

    ###--------------------------------Form8-1--------------------------------------------------
    ws1 = wb.active
    ws1.title = 'Form8-1'

    #output_file_path = f"/home/Mega_App/media/{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    #output_file_path = os.path.join(output_dir,f"{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
    output_file_path = f"media/{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    #output_file_path = os.path.join(settings.MEDIA_ROOT, output_file_name)
    fs = FileSystemStorage(location=settings.MEDIA_ROOT)

    row_num = 2
    cell = ws1.cell(row=row_num, column=2, value="Forma № 8-1")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers1 = [
        ["Baza sığorta haqqı və baza sığorta haqqının katastrofik risk təminatına düşən hissəsi"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]
    cell = ws1.cell(row=1, column=len(form1_a.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws1, row_num, headers1, form1_a)
    ws1.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    cell = ws1.cell(row=row_num, column=2, value=f"Sığorta sinifi  {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form1_a.columns.tolist()
    row_num += 1
    row_num = add_table(ws1, row_num, columns1, form1_a)

    row_num += 1
    columns2 = form1_b.columns.tolist()
    row_num += 1
    row_num = add_table(ws1, row_num, columns2, form1_b)

    signatures = [
        ("İdarə heyətinin sədri ____________________________       _____________",
         "(s.a.a)                                     (imza)"),
        ("Baş mühasib _______________________       _____________",
         "(s.a.a)                                      (imza)"),
        ("M.Y.", "")
    ]
    row_num += 2
    for signature, subtext in signatures:
        ws1.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws1.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws1.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws1.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [30, 25, 20, 40, 25, 40, 50, 45]
    for i, width in enumerate(column_widths, 2):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws1.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###-------------------Form8-2--------------------------------------

    ws2 = wb.create_sheet(title='Form8-2')

    row_num = 2
    cell = ws2.cell(row=row_num, column=2, value="Forma № 8-2")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers2 = [
        ["Baza sığorta haqqı və baza sığorta haqqının katastrofik risk təminatına düşən hissəsi"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]
    cell = ws2.cell(row=1, column=len(form2_a.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws2, row_num, headers2, form2_a)
    ws2.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    cell = ws2.cell(row=row_num, column=2, value=f"Sığorta sinifi  {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form2_a.columns.tolist()
    row_num += 1
    row_num = add_table(ws2, row_num, columns1, form2_a)

    row_num += 1
    columns2 = form2_b.columns.tolist()
    row_num += 1
    row_num = add_table(ws2, row_num, columns2, form_2yekun[0])

    row_num += 2
    for signature, subtext in signatures:
        ws2.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws2.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws2.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws2.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [30, 25, 20, 40, 25, 40, 50, 45]
    for i, width in enumerate(column_widths, 2):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws2.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###-------------------Form8-3--------------------------------------

    ws3 = wb.create_sheet(title='Form8-3')

    row_num = 2
    cell = ws3.cell(row=row_num, column=2, value="Forma № 8-3")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers3 = [
        ["Bildirilmiş, lakin tənzimlənməmiş zərərlər ehtiyatı"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]
    cell = ws3.cell(row=1, column=len(form_3.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws3, row_num, headers3, form_3)
    ws3.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    cell = ws3.cell(row=row_num, column=2, value=f"Sığorta sinifi  {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_3.columns.tolist()
    row_num += 1
    ws3.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws3.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill

    row_num += 1
    ws3.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    row_num += 1
    for index, row in form_3.iterrows():
        ws3.cell(row=row_num, column=1, value=f"A{index + 1}")
        for col_index, value in enumerate(row, start=2):
            ws3.cell(row=row_num, column=col_index, value=value)
        row_num += 1
    row_num += 2
    for signature, subtext in signatures:
        ws3.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws3.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws3.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws3.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [40, 30, 30, 30, 40, 35, 35]
    for i, width in enumerate(column_widths, 2):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws3.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###-------------------Form8-4--------------------------

    ws4 = wb.create_sheet("Form8-4")

    row_num = 2
    cell = ws4.cell(row=row_num, column=2, value="Forma № 8-4")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers4 = [
        ["Baza təkrarsığorta haqqı və baza təkrarsığorta haqqının katastrofik risk təminatına düşən hissəsi"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws4.cell(row=1, column=len(form_4[0].columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws4, row_num, headers4, form_4[0])
    ws4.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    cell = ws4.cell(row=row_num, column=2, value=f"Sığorta sinifi  {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_4[0].columns.tolist()
    row_num += 1
    ws4.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws4.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1
    ws4.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws4.cell(row=row_num, column=3, value="I qrup təkrarsığortaçılar")
    row_num += 1
    for index, row in form_4[0].iterrows():
        for col_index, value in enumerate(row, start=2):
            ws4.cell(row=row_num, column=col_index, value=value)
        row_num += 1
    row_num += 1

    ws4.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws4.cell(row=row_num, column=3, value="II qrup təkrarsığortaçılar")
    row_num += 1
    for index, row in form_4[1].iterrows():
        for col_index, value in enumerate(row, start=2):
            ws4.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    row_num += 1
    ws4.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws4.cell(row=row_num, column=3, value="III qrup təkrarsığortaçılar")
    row_num += 1
    for index, row in form_4[2].iterrows():
        for col_index, value in enumerate(row, start=2):
            ws4.cell(row=row_num, column=col_index, value=value)
        row_num += 1
    row_num += 1
    ws4.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws4.cell(row=row_num, column=3, value="IV qrup təkrarsığortaçılar")
    row_num += 1
    for index, row in form_4yekun[0].iterrows():
        for col_index, value in enumerate(row, start=2):
            ws4.cell(row=row_num, column=col_index, value=value)
        row_num += 1
    row_num += 1

    row_num += 2
    for signature, subtext in signatures:
        ws4.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws4.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws4.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws4.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [40, 30, 30, 30, 40, 60, 50, 65]
    for i, width in enumerate(column_widths, 2):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws4.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###------------------Form8-5--------------------

    ws5 = wb.create_sheet("Form8-5")

    row_num = 2
    cell = ws5.cell(row=row_num, column=2, value="Forma № 8-5")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers5 = [
        [
            "Qazanılmamış sığorta haqları ehtiyatının baza hissəsində təkrarsığortaçıların payı və katastrofik risk təminatına düşən hissəsi"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws5.cell(row=1, column=len(form_5[0].columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws5, row_num, headers5, form_5[0])
    ws5.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    cell = ws5.cell(row=row_num, column=2, value=f"Sığorta sinifi {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_5[0].columns.tolist()
    row_num += 1
    ws5.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws5.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1
    ws5.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws5.cell(row=row_num, column=3, value="I qrup təkrarsığortaçılar")
    row_num += 1
    for index, row in form_5[0].iterrows():
        for col_index, value in enumerate(row, start=2):
            ws5.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    ws5.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws5.cell(row=row_num, column=3, value="II qrup təkrarsığortaçılar")
    row_num += 1
    for index, row in form_5[1].iterrows():
        for col_index, value in enumerate(row, start=2):
            ws5.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    ws5.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws5.cell(row=row_num, column=3, value="III qrup təkrarsığortaçılar")
    row_num += 1
    for index, row in form_5[2].iterrows():
        for col_index, value in enumerate(row, start=2):
            ws5.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    row_num += 1
    ws5.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws5.cell(row=row_num, column=3, value="IV qrup təkrarsığortaçılar")
    row_num += 1
    for index, row in form_5yekun[0].iterrows():
        for col_index, value in enumerate(row, start=2):
            ws5.cell(row=row_num, column=col_index, value=value)
        row_num += 1
    row_num += 2
    for signature, subtext in signatures:
        ws5.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws5.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws5.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws5.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [40, 30, 30, 30, 40, 60, 50, 65]
    for i, width in enumerate(column_widths, 2):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws5.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###------------------------Form8-6-------------------------
    ws6 = wb.create_sheet("Form8-6")

    row_num = 2
    cell = ws6.cell(row=row_num, column=2, value="Forma № 8-6")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers6 = [
        ["Qazanılmamış sığorta haqları ehtiyatının əlavə hissəsi"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws6.cell(row=1, column=len(form_6.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws6, row_num, headers6, form_6)
    cell = ws6.cell(row=row_num, column=2, value=f"Sığorta sinifi  {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_6.columns.tolist()
    row_num += 1
    ws6.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws6.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1

    for index, row in form_6.iterrows():
        for col_index, value in enumerate(row, start=2):
            ws6.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    row_num += 2
    for signature, subtext in signatures:
        ws6.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws6.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws6.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws6.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [25, 30, 25, 30, 40, 60, 50, 65]
    for i, width in enumerate(column_widths, 2):
        ws6.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws6.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###------------------------------Form8_7---------------------

    ws7 = wb.create_sheet("Form8-7")

    row_num = 2
    cell = ws7.cell(row=row_num, column=2, value="Forma № 8-7")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers7 = [
        [" Qazanılmış məcmu sığorta (təkrarsığorta) haqları və katastrofik risk təminatına düşən hissəsi"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws7.cell(row=1, column=len(form_7.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws7, row_num, headers7, form_7)
    cell = ws7.cell(row=row_num, column=2, value=f"Sığorta sinfi {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_7.columns.tolist()
    row_num += 1
    ws7.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws7.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1

    for index, row in form_7.iterrows():
        ws7.cell(row=row_num, column=1, value=f"A{index + 1}")

        for col_index, value in enumerate(row, start=2):
            ws7.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    row_num += 2
    for signature, subtext in signatures:
        ws7.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws7.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws7.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws7.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [35, 20, 25, 30, 30, 30, 40, 30, 45, 45]
    for i, width in enumerate(column_widths, 2):
        ws7.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws7.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###---------------------------Form8_8--------------------

    ws8 = wb.create_sheet("Form8-8")

    row_num = 2
    cell = ws8.cell(row=row_num, column=2, value="Forma № 8-8")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers8 = [
        ["Üçbucaq metodu ilə hesablanan baş vermiş, lakin bildirilməmiş zərərlər ehtiyatı"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws8.cell(row=1, column=len(form_8.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws8, row_num, headers8, form_8)
    ws8.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    cell = ws8.cell(row=row_num, column=2, value=f"Sığorta sinifi {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_8.columns.tolist()
    row_num += 1
    ws8.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 3):
        cell = ws8.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1
    row_num2 = row_num
    for index, row in form_8.iterrows():
        if isinstance(index, pd.Timestamp):
            index = index.to_pydatetime().date().toordinal()
        ws8.cell(row=row_num, column=1, value=f"A{index + 1}")

        for col_index, value in enumerate(row, start=3):
            ws8.cell(row=row_num, column=col_index, value=value)

        row_num += 1

    ws8.merge_cells(start_row=row_num2, start_column=2, end_row=row_num - 8, end_column=2)
    cell = ws8.cell(row=row_num2, column=2, value="Zərərlərin baş verdiyi rüblər")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.fill = fill

    texts = [
        "Zərərlərin baş verdiyi rüblər ərzində baş vermiş və zərərlərin ödənildiyi (inkişafı) rübün sonunadək ödənilmiş zərərlərin məcmu məbləği",
        "Zərərlərin baş verdiyi rüblər (axırıncı rüb istisna olmaqla) ərzində baş vermiş və zərərlərin ödənildiyi (inkişafı) rübün sonunadək ödənilmiş zərərlərin məcmu məbləği",
        "Zərərlərin inkişafı əmsalları",
        "Zərərlərin inkişafı amilləri",
        "Gecikmə amilləri",
        "Baş vermiş, lakin bildirilməmiş zərərlərin məbləği",
        "Üçbucaq metodu ilə hesblanmış BVBZE"
    ]
    for i, text in enumerate(texts, start=row_num - 7):
        cell = ws8.cell(row=i, column=2, value=text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill

    for row in range(row_num - 7, row_num):
        ws8.row_dimensions[row].height = 15
    row_num += 2
    for signature, subtext in signatures:
        ws8.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws8.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws8.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws8.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    if len(columns1) <21:
        column_widths = [25, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 30, 30, 30, 30, 30, 30, 30]
    else:
        column_widths = [25, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10,10, 10, 10, 10, 10, 10, 
                         10, 10, 30, 30, 30, 30, 30, 30, 30]

    for i, width in enumerate(column_widths, 2):
        ws8.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws8.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###---------------------------Form8_9--------------------

    ws9 = wb.create_sheet("Form8-9")

    row_num = 2
    cell = ws9.cell(row=row_num, column=2, value="Forma № 8-9")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers9 = [
        ["Baş vermiş, lakin bildirilməmiş zərərlər ehtiyatı"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws9.cell(row=1, column=len(form_9.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws9, row_num, headers9, form_9)
    cell = ws9.cell(row=row_num, column=2, value=f"Sığorta sinfi {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_9.columns.tolist()
    row_num += 1
    ws9.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws9.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1

    for index, row in form_9.iterrows():
        ws9.cell(row=row_num, column=1, value=f"A{index + 1}")

        for col_index, value in enumerate(row, start=2):
            ws9.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    row_num += 2
    for signature, subtext in signatures:
        ws9.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws9.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws9.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws9.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [25, 30, 25, 30]
    for i, width in enumerate(column_widths, 2):
        ws9.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws9.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###---------------------------Form8_10-----------------------

    ws10 = wb.create_sheet("Form8-10")

    row_num = 2
    cell = ws10.cell(row=row_num, column=2, value="Forma № 8-10")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers10 = [
        ["Qazanılmış məcmu təkrarsığorta haqları və katastrofik risk təminatına düşən hissəsi"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws10.cell(row=1, column=len(form_10.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws10, row_num, headers10, form_10)
    cell = ws10.cell(row=row_num, column=2, value=f"Sığorta sinifi {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_10.columns.tolist()
    row_num += 1
    ws10.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws10.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1

    for index, row in form_10.iterrows():
        ws10.cell(row=row_num, column=1, value=f"A{index + 1}")

        for col_index, value in enumerate(row, start=2):
            ws10.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    row_num += 2
    for signature, subtext in signatures:
        ws10.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws10.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws10.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws7.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [35, 20, 25, 30, 30, 30, 40, 30, 45, 45]
    for i, width in enumerate(column_widths, 2):
        ws10.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws10.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###----------------------Form8_11-------------------------------

    ws11 = wb.create_sheet(title='Form8-11')

    row_num = 2
    cell = ws11.cell(row=row_num, column=2, value="Forma № 8-11")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers11 = [
        ["Bildirilmiş, lakin tənzimlənməmiş zərərlər ehtiyatında təkrarsığortaçıların payı"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]
    cell = ws11.cell(row=1, column=len(form_11.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws11, row_num, headers11, form_11)
    ws11.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    cell = ws11.cell(row=row_num, column=2, value=f"Sığorta sinifi {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_11.columns.tolist()
    row_num += 1
    ws11.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws11.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill

    row_num += 1
    ws11.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    row_num += 1
    for index, row in form_11.iterrows():
        ws11.cell(row=row_num, column=1, value=f"A{index + 1}")
        for col_index, value in enumerate(row, start=2):
            ws11.cell(row=row_num, column=col_index, value=value)
        row_num += 1
    row_num += 2
    for signature, subtext in signatures:
        ws11.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws11.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws11.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws11.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [40, 30, 30, 30, 40, 35, 35]
    for i, width in enumerate(column_widths, 2):
        ws11.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws11.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###---------------------Form8-12--------------------------------
    ws12 = wb.create_sheet("Form8-12")

    row_num = 2
    cell = ws12.cell(row=row_num, column=2, value="Forma № 8-12")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers12 = [
        [
            "Üçbucaq metodu ilə hesablanan baş vermiş, lakin bildirilməmiş zərərlər ehtiyatında təkrarsığortaçıların payı"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws12.cell(row=1, column=len(form_12.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws12, row_num, headers12, form_12)
    ws12.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
    cell = ws12.cell(row=row_num, column=2, value=f"Sığorta sinifi  {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_12.columns.tolist()
    row_num += 1
    ws12.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 3):
        cell = ws12.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1
    row_num2 = row_num
    for index, row in form_12.iterrows():

        if isinstance(index, pd.Timestamp):
            index = index.to_pydatetime().date().toordinal()
        ws12.cell(row=row_num, column=1, value=f"A{index + 1}")

        for col_index, value in enumerate(row, start=3):
            ws12.cell(row=row_num, column=col_index, value=value)

        row_num += 1

    ws12.merge_cells(start_row=row_num2, start_column=2, end_row=row_num - 8, end_column=2)
    cell = ws12.cell(row=row_num2, column=2, value="Zərərlərin baş verdiyi rüblər")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = fill

    texts = [
        "Zərərlərin baş verdiyi rüblər ərzində baş vermiş və zərərlərin ödənildiyi (inkişafı) rübün sonunadək ödənilmiş zərərlərin məcmu məbləği",
        "Zərərlərin baş verdiyi rüblər (axırıncı rüb istisna olmaqla) ərzində baş vermiş və zərərlərin ödənildiyi (inkişafı) rübün sonunadək ödənilmiş zərərlərin məcmu məbləği",
        "Zərərlərin inkişafı əmsalları",
        "Zərərlərin inkişafı amilləri",
        "Gecikmə amilləri",
        "Baş vermiş, lakin bildirilməmiş zərərlərin məbləği",
        "Üçbucaq metodu ilə hesblanmış BVBZE"
    ]
    for i, text in enumerate(texts, start=row_num - 7):
        cell = ws12.cell(row=i, column=2, value=text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill

    for row in range(row_num - 7, row_num):
        ws12.row_dimensions[row].height = 15

    row_num += 2
    for signature, subtext in signatures:
        ws12.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws12.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws12.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws12.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    if len(columns1) <21:
        column_widths = [25, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 30, 30, 30, 30, 30, 30, 30]
    else:
        column_widths = [25, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10,10, 10, 10, 10, 10, 10, 
                         10, 10, 30, 30, 30, 30, 30, 30, 30]
    for i, width in enumerate(column_widths, 2):
        ws12.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws12.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ###-------------------Form8-13-----------------------

    ws13 = wb.create_sheet("Form8-13")

    row_num = 2
    cell = ws13.cell(row=row_num, column=2, value="Forma № 8-13")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    headers9 = [
        ["Baş vermiş, lakin bildirilməmiş zərərlər ehtiyatı"],
        ["__________________________________________"],
        ["(sığortaçının (təkrarsığortaçının) adı)"],
        [f"{formatted_date}-cü il  tarixə"]
    ]

    cell = ws13.cell(row=1, column=len(form_13.columns) + 1, value="""«Həyat sığortası və qeyri-həyat sığortası
    üzrə sığorta ehtiyatlarının formalaşdırılması
    Qaydası»na Əlavə № 3""")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = write_headers(ws13, row_num, headers9, form_13)
    cell = ws13.cell(row=row_num, column=2, value=f"Sığorta sinfi {class_of_insurance}")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")
    row_num += 1

    columns1 = form_13.columns.tolist()
    row_num += 1
    ws13.cell(row=row_num, column=1, value='A')
    fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for col_num, col_data in enumerate(columns1, 2):
        cell = ws13.cell(row=row_num, column=col_num, value=col_data)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = fill
    row_num += 1

    for index, row in form_13.iterrows():
        ws13.cell(row=row_num, column=1, value=f"A{index + 1}")

        for col_index, value in enumerate(row, start=2):
            ws13.cell(row=row_num, column=col_index, value=value)
        row_num += 1

    row_num += 2
    for signature, subtext in signatures:
        ws13.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        ws13.cell(row=row_num, column=2, value=signature)
        row_num += 1
        if subtext:
            ws13.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
            ws13.cell(row=row_num, column=2, value=subtext)
            row_num += 1

    column_widths = [25, 30, 25, 30]
    for i, width in enumerate(column_widths, 2):
        ws13.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws13.iter_rows(min_row=1, max_row=row_num):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(output_file_path)
    #fs = FileSystemStorage(location=settings.MEDIA_ROOT)
    #fs.save(output_file_name, output_file_name)
    return output_file_path
